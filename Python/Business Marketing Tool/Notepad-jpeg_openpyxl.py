import sys
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QLabel
from PyQt5.QtGui import QPixmap, QPainter
from PyQt5.QtCore import Qt
from tempfile import NamedTemporaryFile
import pandas as pd
import openpyxl
import time
import webbrowser
import pyautogui
import os
from glob import glob
from urllib.parse import urlparse
from datetime import datetime
import subprocess
from PIL import Image, ImageDraw, ImageFont

"""
A hard-coded name needs to be represented in the cell that divides the active informants from the non-active informants for this script to work successfully.
"""

non_blue_colors = ['FFFFFF00', 'FF7030A0', 'FF00B050']
# blue_color = 'FF0070C0'
red_color = ['FFC00000', 'FF0000']

sharepoint_link = 'https://companyname.sharepoint/Excel.xlsx'

webbrowser.open_new(bfu_sharepoint_link)
time.sleep(1.5)
pyautogui.hotkey('win', '7')

# To obtain the most recent excel file in downloads
downloads_path = os.path.join(os.path.expanduser('~'), 'Downloads')
downloaded_excel_files = glob(os.path.join(downloads_path, '*.xlsx'))
INFORMANT_LIST = max(downloaded_excel_files, key=os.path.getctime)

wb = openpyxl.load_workbook(INFORMANT_LIST, data_only=True)
sheet = wb.worksheets[0]

informants_on_leave = []
informants_in_training = []

for cell in sheet['A']:
    if cell.value != 'INFORMANT':
        cell_color = cell.fill.start_color.rgb

        if cell.value == "HARCODED NAME":
            break

        if cell_color not in non_blue_colors: # This logic handles colors that aren't excel's standard primary colors
            names = cell.value
            employee_ids = sheet.cell(row=cell.row, column=2).value
            informants_on_leave.append((names, employee_ids))

        if cell_color in red_color:
            names = cell.value
            employee_ids = sheet.cell(row=cell.row, column=2).value
            informants_in_training.append((names, employee_ids))

BAT_informants_List = 'C:\\Users\\ExcelFile.xlsx'
df2 = pd.read_excel(BAT_informants_List, sheet_name='Sheet1')

ignore_names = ['IGNORE NAME 1', "IGNORE NAME 2", "IGNORE NAME 3"]
df1 = pd.read_excel(INFORMANT_LIST, sheet_name='informants', usecols=['INFORMANT', 'ID', 'B', 'C', 'D', 'E'])

new_informants = []
for name in df1['INFORMANT']:
    if name == "HARCODED NAME":
        break
    if name not in ignore_names:
        new_informants.append(name)

name_columns = ['Name', 'Name.1', 'Name.2', 'Name.3']

def clean_name(name):
    name = name.replace(' (OL)', '').replace(' (I.T)', '').strip().replace(' (AWHOL)', '').strip()
    return name

BAT_Security_Experts = []

for col_name in name_columns:
    Table_Contained_Names = [clean_name(name) for name in df2[col_name].dropna().unique().tolist()]
    BAT_Security_Experts.extend(Table_Contained_Names)

new_seller_names = set(new_informants) - set(BAT_Security_Experts)

clean_new_informants = {x for x in new_seller_names if x == x}

new_seller_names_ids = [(row.INFORMANT, row.ID) for index, row in df1[df1['INFORMANT'].isin(clean_new_informants)].iterrows()]

dismissed_seller_names = set(BAT_Security_Experts) - set(new_informants)

leavers_names_ids = [(row.INFORMANT, row.ID) for index, row in df1[df1['INFORMANT'].isin(dismissed_seller_names)].iterrows()]

app = QApplication(sys.argv)

widget = QWidget()
layout = QVBoxLayout(widget)

widget.setStyleSheet("""
    QWidget {
        background-color: #AFABAB; /* Dark blue background */
        color: #C00000; /* red text */
        font-size = 12pt;
    } 
""")

def add_section_to_layout(title, items):
    if not items:
        placeholder = "No Data Available"
        layout.addWidget(QLabel(f'<b>{title}</b>'))
        layout.addWidget(QLabel(placeholder))
        layout.addWidget(QLabel(''))
    else:
        layout.addWidget(QLabel(f'<b>{title}</b>'))
        for item in items:
            if isinstance(item, tuple) and len(item) == 2:
                name, id = item
                item_st = f"{name} - {id}"
            else:
                item_st = str(item)

            layout.addWidget(QLabel(item_st))
        layout.addWidget(QLabel(''))

add_section_to_layout("New informants to be added:", new_seller_names_ids)
add_section_to_layout("informants to be removed", leavers_names_ids)
add_section_to_layout("Informants on leave:", informants_on_leave)
add_section_to_layout("Informants in training:", informants_in_training)

widget.adjustSize()

pixmap = QPixmap(widget.size())
painter = QPainter(pixmap)
widget.render(painter)
painter.end()

with NamedTemporaryFile(delete=False, suffix='.png') as tmpfile:
    pixmap.save(tmpfile.name)
    print(f'Saved snapshot to {tmpfile.name}')

if sys.platform.startswith('win'):
    subprocess.Popen(['explorer', tmpfile.name])
else:
    print("OS is not native to Windows")

# Extract data from SharePoint excel workbook for new informants
new_informants_data = []
if not df1.empty:
    for row in df1[df1['INFORMANT'].isin(clean_new_informants)].itertuples(index=False, name=None):
        new_informants_data.append((row[0], row[2], row[3], "", row[4], row[5]))  # Adding empty string between columns C and D

    # Create notepad file
    notepad_filename = os.path.join(downloads_path, 'New_Informants_Data.txt')
    with open(notepad_filename, 'w') as file:
        for data_row in new_informants_data:
            line = '\t'.join(str(item) for item in data_row)
            file.write(line + '\n\n')

    print(f'Saved new informants data to {notepad_filename}')

    if sys.platform.startswith('win'):
        subprocess.Popen(['notepad', notepad_filename])
    else:
        print("OS is not native to Windows")
else:
    print("No new informants data available")
